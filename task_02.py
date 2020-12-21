import openpyxl
import pandas as pd

# вариант 3

result = openpyxl.Workbook()  # чтение файлов
items = pd.read_csv('items.csv', sep=';', encoding='utf8')
persons = pd.read_csv('persons.csv', sep=';', encoding='utf8')

result.create_sheet('Persons', 0)  # создание листов
result.create_sheet('Items', 1)
del result['Sheet']


def insert_names(name: str, my_list):  # заполнение страниц данными
    sheet = result[f'{name}']
    i = 1
    for row in my_list:
        print(row)
        sheet.cell(row=1, column=i).value = row
        i += 1


def insert_data(name: str, my_list):
    sheet = result[f'{name}']
    i = 2
    j = 1
    for row in my_list:
        for value in my_list[f'{row}']:
            sheet.cell(row=i, column=j).value = value
            i += 1
        j += 1
        i = 2


insert_names('Items', items)
insert_names('Persons', persons)
insert_data('Items', items)
insert_data('Persons', persons)

result.save('persons_and_items.xlsx')
result.close()
